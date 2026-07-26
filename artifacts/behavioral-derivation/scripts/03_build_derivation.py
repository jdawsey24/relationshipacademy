#!/usr/bin/env python3
"""Step 3 — merge derivation outputs + run mechanical global-consistency checks + build the workbook.
Reads ../generated/{bank,behaviors,cluster_meta,fw_competencies,qa_agent}.json + ../generated/enrichment/cluster_*.json;
writes ../outputs/RLC_Behavioral_Derivation.xlsx (Derivation | Play Synthesis | Human Review | Global QA).
"""
"""Merge Script-2 derivation outputs, run mechanical global-consistency checks, build the workbook.
Sheets: Derivation | Play Synthesis | Human Review | Global QA.
Reads out2/cluster_*_behaviors.json + _plays.json; optional qa_agent.json (consistency-agent judgment findings)."""
import json, glob, os, sys, collections, re
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEN = os.path.join(ROOT,"generated")
OUT = os.path.join(ROOT,"outputs"); os.makedirs(OUT,exist_ok=True)
SP = GEN
def S(v): return "; ".join(map(str, v)) if isinstance(v, (list, tuple)) else ("" if v is None else str(v))
def IDS(v): return ", ".join(map(str, v)) if isinstance(v, (list, tuple)) else ("" if v is None else str(v))
def as_list(v): return v if isinstance(v, list) else ([x.strip() for x in str(v).split(",") if x.strip()] if v else [])
bank = {b["sid"]: b for b in json.load(open(f"{SP}/bank.json"))}
behaviors = {b["sid"]: b for b in json.load(open(f"{SP}/behaviors.json"))}   # the 81 targets
meta = json.load(open(f"{SP}/cluster_meta.json")); cname = {int(k): v["name"] for k, v in meta.items()}
fwc = json.load(open(f"{SP}/fw_competencies.json"))
# valid competency names by (phase, domain) and globally
valid_comp = collections.defaultdict(set)
for c in fwc: valid_comp[c["phase"]].add((c["domain"].strip(), c["name"].strip()))
all_comp_names = {c["name"].strip() for c in fwc}
SPECIAL_COMP = {"None available (source gap)", "Task-supported; no competency cleanly maps", ""}

# ---- load derivation ----
der = {}
for fp in sorted(glob.glob(f"{GEN}/enrichment/cluster_*_behaviors.json")):
    for r in json.load(open(fp)): der[r["sid"]] = r
plays = []
for fp in sorted(glob.glob(f"{GEN}/enrichment/cluster_*_plays.json")):
    cid = int(re.search(r"cluster_(\d+)_plays", fp).group(1))
    for p in json.load(open(fp)): plays.append({"_cluster": cid, **p})
for p in plays:
    p["behavioral_source_ids"] = as_list(p.get("behavioral_source_ids"))
    p["contextual_source_ids"] = as_list(p.get("contextual_source_ids"))

# ---- validate coverage ----
missing = [s for s in behaviors if s not in der]
extra = [s for s in der if s not in behaviors]
print(f"target behaviors={len(behaviors)} derived={len(der)} missing={len(missing)} extra={len(extra)} plays={len(plays)}")
if missing: print("MISSING:", missing[:20])
if extra: print("EXTRA:", extra[:20])
if missing or extra:
    print("!! incomplete — not building"); sys.exit(1)

# ---- mechanical consistency checks ----
findings = {}
# (8) Insufficient-Evidence sid used as behavioral evidence
insuff = {s for s in der if der[s]["maintaining_role"] == "Insufficient Evidence"}
bad_evidence = [(p["play_name"], sid) for p in plays for sid in p.get("behavioral_source_ids", []) if sid in insuff]
findings["insufficient_as_behavioral_evidence"] = bad_evidence
# (3) invented/unknown competency labels (behaviors + plays)
def check_comp(label):
    if label in SPECIAL_COMP: return True
    # accept "Domain -> Name" or names containing a known competency name
    names = re.split(r"[;,]", label)
    ok = False
    for n in names:
        m = re.search(r"→\s*(.+)$", n.strip()) or re.search(r"->\s*(.+)$", n.strip())
        nm = (m.group(1) if m else n).strip()
        for known in all_comp_names:
            if known and known.lower() in nm.lower(): ok = True
    return ok
unknown_comp = sorted({der[s]["relevant_rlc_competency"] for s in der
                       if der[s]["relevant_rlc_competency"] and not check_comp(der[s]["relevant_rlc_competency"])})
findings["unknown_competency_labels"] = unknown_comp
# (9/10) plays per cluster; clusters with no plays
plays_per_cluster = collections.Counter(p["_cluster"] for p in plays)
clusters_with_behaviors = sorted({behaviors[s]["primary_ec"] for s in behaviors})
findings["plays_per_cluster"] = {c: plays_per_cluster.get(c, 0) for c in clusters_with_behaviors}
findings["clusters_no_plays"] = [c for c in clusters_with_behaviors if plays_per_cluster.get(c, 0) == 0]
findings["clusters_many_plays"] = [c for c in clusters_with_behaviors if plays_per_cluster.get(c, 0) >= 4]
# (1) duplicate statement texts getting different maintaining_role/function across clusters
bytext = collections.defaultdict(list)
for s in behaviors: bytext[bank[s]["statement"].strip()].append(s)
dup_conflicts = []
for txt, sids in bytext.items():
    if len(sids) > 1:
        roles = {der[s]["maintaining_role"] for s in sids}
        if len(roles) > 1: dup_conflicts.append((txt, [(s, der[s]["maintaining_role"]) for s in sids]))
findings["duplicate_text_role_conflicts"] = dup_conflicts
# QA rollups
roles = collections.Counter(der[s]["maintaining_role"] for s in der)
context_dep = [s for s in der if der[s]["maintaining_role"] == "Context Dependent"]
task_only = [s for s in der if der[s]["relevant_rlc_competency"] in ("Task-supported; no competency cleanly maps", "None available (source gap)")]
flagged_beh = [s for s in der if der[s]["review_flag"]]
flagged_plays = [p for p in plays if p.get("review_flag")]
consolidation = {c: {"behaviors": sum(1 for s in behaviors if behaviors[s]["primary_ec"] == c and der[s]["maintaining_role"] in ("Likely Maintains Pattern", "Context Dependent")),
                     "plays": plays_per_cluster.get(c, 0)} for c in clusters_with_behaviors}
qa_agent = json.load(open(f"{SP}/qa_agent.json")) if os.path.exists(f"{SP}/qa_agent.json") else None

# ================= build workbook =================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
wb = openpyxl.Workbook()
hf = PatternFill("solid", fgColor="1C3557"); hfont = Font(bold=True, color="FFFFFF")
def hdr(ws, n):
    for j in range(1, n + 1):
        c = ws.cell(1, j); c.fill = hf; c.font = hfont; c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(n)}{ws.max_row}"

# --- Derivation ---
DCOLS = ["Statement_ID","Experience_Cluster","RLC_Phase","Original_Behavior","Trigger_or_Context","Maintaining_Role",
         "Short_Term_Function","Developmental_Cost","Relevant_Developmental_Task","Relevant_RLC_Competency",
         "Adaptive_Alternative_1","Adaptive_Alternative_2","What_New_Response_Supports","Expected_Discomfort_or_Cost",
         "Context_Conditions","Confidence","Review_Flag"]
ws = wb.active; ws.title = "Derivation"; ws.append(DCOLS)
for s in sorted(der, key=lambda x: (behaviors[x]["primary_ec"], x)):
    d = der[s]; b = behaviors[s]
    ws.append([s, f'{b["primary_ec"]} — {cname.get(b["primary_ec"],"?")}', b["phase"], bank[s]["statement"],
               S(d.get("trigger_or_context")), S(d.get("maintaining_role")), S(d.get("short_term_function")), S(d.get("developmental_cost")),
               S(d.get("relevant_developmental_task")), S(d.get("relevant_rlc_competency")), S(d.get("adaptive_alternative_1")),
               S(d.get("adaptive_alternative_2")), S(d.get("what_new_response_supports")), S(d.get("expected_discomfort_or_cost")),
               S(d.get("context_conditions")), S(d.get("confidence")), "YES" if d.get("review_flag") else "NO"])
hdr(ws, len(DCOLS))
for j, w in enumerate([12,26,13,34,34,22,40,36,20,30,44,40,36,32,34,10,9], 1): ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

# --- Play Synthesis ---
PCOLS = ["Play_ID","Play_Name","Experience_Cluster","RLC_Phase","Relevant_Developmental_Task","Behavioral_Target","Trigger",
         "Old_Play","Function_of_Old_Play","New_Play","What_to_Expect","Why_Supports_Task","Supporting_RLC_Competencies",
         "Behavioral_Source_IDs","Contextual_Source_IDs","Maintaining_Behaviors_Consolidated","Context_Limitations","Confidence","Review_Flag","Review_Reason"]
ps = wb.create_sheet("Play Synthesis"); ps.append(PCOLS)
plays_sorted = sorted(plays, key=lambda p: p["_cluster"])
for i, p in enumerate(plays_sorted, 1):
    cid = p["_cluster"]; ph = next((behaviors[s]["phase"] for s in behaviors if behaviors[s]["primary_ec"] == cid), "")
    ps.append([f"PLAY-{i:03d}", S(p.get("play_name")), f'{cid} — {cname.get(cid,"?")}', ph, "",
               S(p.get("behavioral_target")), S(p.get("trigger")), S(p.get("old_play")), S(p.get("function_of_old_play")),
               S(p.get("new_play")), S(p.get("what_to_expect")), S(p.get("why_supports_task")), S(p.get("supporting_competencies")),
               IDS(p.get("behavioral_source_ids")), IDS(p.get("contextual_source_ids")),
               S(p.get("maintaining_behaviors_consolidated")), S(p.get("context_limitations")), S(p.get("confidence")),
               "YES" if p.get("review_flag") else "NO", S(p.get("review_reason"))])
# fill Relevant_Developmental_Task from a behavior in the cluster
for row in range(2, ps.max_row+1):
    cid = int(str(ps.cell(row,3).value).split(" — ")[0])
    task = next((der[s]["relevant_developmental_task"] for s in der if behaviors[s]["primary_ec"]==cid), "")
    ps.cell(row,5).value = task
hdr(ps, len(PCOLS))
for j, w in enumerate([9,30,26,13,22,34,30,34,34,44,36,36,34,20,18,34,34,10,9,34], 1): ps.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

# --- Human Review ---
HCOLS = ["Item_Type","ID","Experience_Cluster","RLC_Phase","Maintaining_Role","Name_or_Behavior","Review_Reason","Confidence"]
hr = wb.create_sheet("Human Review"); hr.append(HCOLS)
for s in sorted(flagged_beh, key=lambda x:(behaviors[x]["primary_ec"],x)):
    d=der[s]; b=behaviors[s]
    hr.append(["Behavior", s, f'{b["primary_ec"]} — {cname.get(b["primary_ec"],"?")}', b["phase"], d["maintaining_role"], bank[s]["statement"], d.get("review_reason",""), d["confidence"]])
for i,p in enumerate(plays_sorted,1):
    if p.get("review_flag"):
        cid=p["_cluster"]
        hr.append(["Play", f"PLAY-{i:03d}", f'{cid} — {cname.get(cid,"?")}', "", "", p.get("play_name",""), p.get("review_reason",""), p.get("confidence","")])
hdr(hr, len(HCOLS))
for j,w in enumerate([11,11,26,13,22,44,50,10],1): hr.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w

# --- Global QA ---
qa = wb.create_sheet("Global QA"); qa.cell(1,1,"Script 2 — Global Consistency & QA").font=Font(bold=True,size=13)
r=[3]
def line(a,b=""):
    qa.cell(r[0],1,a); qa.cell(r[0],2,b); r[0]+=1
line("Total behaviors analyzed", len(der)); line("Total candidate Plays", len(plays))
line("Maintaining_Role counts",""); [line(f"   {k}", v) for k,v in roles.most_common()]
line("")
line("QA-1 Context-Dependent behaviors (looked maladaptive; ruled conditional)", len(context_dep)); line("   IDs", ", ".join(context_dep))
line("QA-2 Alternatives NOT cleanly traced to a competency (task-only / source gap)", len(task_only)); line("   IDs", ", ".join(task_only))
line("QA-3 Plays flagged for weak/uncertain support", len(flagged_plays))
line("QA-4 Consolidation (behaviors eligible -> plays) per cluster",""); [line(f"   C{c}", f"{v['behaviors']} -> {v['plays']}") for c,v in consolidation.items()]
line("QA-5 Clusters with NO defensible behavioral Play", ", ".join(str(c) for c in findings["clusters_no_plays"]) or "none")
line("QA-6 Items needing human review (behaviors+plays)", f"{len(flagged_beh)} + {len(flagged_plays)}")
line("")
line("GLOBAL CONSISTENCY CHECKS (mechanical)","")
line("C8 Insufficient-Evidence sid used as BEHAVIORAL evidence", str(findings["insufficient_as_behavioral_evidence"]) if findings["insufficient_as_behavioral_evidence"] else "none ✓")
line("C3 Unknown/invented competency labels", ", ".join(findings["unknown_competency_labels"]) or "none ✓")
line("C1 Duplicate-text behaviors with conflicting Maintaining_Role", str([d[0] for d in findings["duplicate_text_role_conflicts"]]) if findings["duplicate_text_role_conflicts"] else "none ✓")
line("C9 Clusters with many Plays (>=4)", ", ".join(str(c) for c in findings["clusters_many_plays"]) or "none")
line("C10 Clusters with no Plays", ", ".join(str(c) for c in findings["clusters_no_plays"]) or "none")
line("")
line("GLOBAL CONSISTENCY (judgment — consistency agent)","")
if qa_agent:
    for k in ("same_behavior_diff_function","inconsistent_alternatives","generic_advice","unsupported_mechanisms","over_consolidated_plays","duplicative_plays_across_clusters"):
        v=qa_agent.get(k,[]); line(f"   {k}", (str(v)[:400] if v else "none reported"))
    if qa_agent.get("notes"): line("   notes", qa_agent["notes"][:400])
else:
    line("   (consistency agent output not yet loaded — qa_agent.json missing)","")
qa.column_dimensions["A"].width=60; qa.column_dimensions["B"].width=80

out=os.path.join(OUT,"RLC_Behavioral_Derivation.xlsx"); wb.save(out)
print("BUILT:", out)
print("roles:", dict(roles), "| flagged beh:", len(flagged_beh), "plays:", len(plays), "flagged plays:", len(flagged_plays))
print("MECH: insuff_as_evidence:", findings["insufficient_as_behavioral_evidence"], "| unknown_comp:", findings["unknown_competency_labels"])
print("plays/cluster:", findings["plays_per_cluster"])
print("no-play clusters:", findings["clusters_no_plays"], "| many-play:", findings["clusters_many_plays"])
