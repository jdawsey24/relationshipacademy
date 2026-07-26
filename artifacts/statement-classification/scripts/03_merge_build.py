#!/usr/bin/env python3
"""Merge canonical bank + per-cluster enrichment into the classification xlsx.
Validates completeness first; refuses to build if any statement is unclassified."""
import json, glob, os, sys, collections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEN = os.path.join(ROOT, "generated")
OUT = os.path.join(ROOT, "outputs"); os.makedirs(OUT, exist_ok=True)
SP = GEN  # bank.json, cluster_meta.json, flags.json live in generated/
bank = json.load(open(f"{SP}/bank.json"))
meta = json.load(open(f"{SP}/cluster_meta.json"))
cname = {int(k): v["name"] for k, v in meta.items()}

# load enrichment
enr = {}
dupes = []
for fp in sorted(glob.glob(f"{GEN}/enrichment/cluster_*.json")):
    arr = json.load(open(fp))
    for r in arr:
        if r["sid"] in enr:
            dupes.append(r["sid"])
        enr[r["sid"]] = r
flags = json.load(open(f"{SP}/flags.json"))  # sid -> {flag_category:[...], flag_reason}

# validate
bank_sids = [b["sid"] for b in bank]
missing = [s for s in bank_sids if s not in enr]
extra = [s for s in enr if s not in set(bank_sids)]
print(f"bank={len(bank)} enriched={len(enr)} missing={len(missing)} extra={len(extra)} dupes={len(dupes)}")
if missing:
    print("MISSING (first 20):", missing[:20]);
if extra:
    print("EXTRA (first 20):", extra[:20])
if missing or extra or dupes:
    print("!! NOT COMPLETE — fix before building."); sys.exit(1)

VALID_TYPES = {"Experience","Emotion","Thought/Belief","Fear","Need","Self-Behavior","Other-Person Behavior","Relational Condition"}
VALID_CONF = {"High","Moderate","Low"}
def clab(cid):
    return "" if cid in (None,"") else f"{int(cid)} — {cname.get(int(cid),'?')}"

rows = []
type_issues=0
for b in bank:
    e = enr[b["sid"]]
    t = e["statement_type"]
    if t not in VALID_TYPES or e["confidence"] not in VALID_CONF: type_issues+=1
    rows.append({
        "Statement_ID": b["sid"],
        "Original_Statement": b["statement"],
        "Statement_Type": t,
        "Primary_Experience_Cluster": clab(b["primary_ec"]),
        "Secondary_Experience_Cluster": clab(e.get("secondary_ec")),
        "RLC_Phase": b["phase"],
        "Behaviorally_Actionable": "YES" if e["behaviorally_actionable"] else "NO",
        "Mapping_Rationale": e["mapping_rationale"],
        "Confidence": e["confidence"],
        "Review_Flag": "YES" if e["review_flag"] else "NO",
        "_primary_id": int(b["primary_ec"]),
    })
if type_issues: print(f"WARN: {type_issues} rows with invalid type/confidence value")
rows.sort(key=lambda r: r["Statement_ID"])

# ---- QA numbers ----
by_type = collections.Counter(r["Statement_Type"] for r in rows)
by_prim = collections.Counter(r["_primary_id"] for r in rows)
n_action = sum(1 for r in rows if r["Behaviorally_Actionable"]=="YES")
n_flag = sum(1 for r in rows if r["Review_Flag"]=="YES")
flag_cat_counts = collections.Counter(c for s in flags for c in flags[s]["flag_category"])
by_sec = collections.Counter(enr[b["sid"]].get("secondary_ec") for b in bank if enr[b["sid"]].get("secondary_ec"))
counts = sorted(by_prim.values())
import statistics
mean=statistics.mean(by_prim.values());
few=[c for c in by_prim if by_prim[c] <= 12]
many=[c for c in by_prim if by_prim[c] >= 60]

# ---- build xlsx ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
wb = openpyxl.Workbook()
COLS = ["Statement_ID","Original_Statement","Statement_Type","Primary_Experience_Cluster",
        "Secondary_Experience_Cluster","RLC_Phase","Behaviorally_Actionable","Mapping_Rationale","Confidence","Review_Flag"]
hdr_fill = PatternFill("solid", fgColor="1C3557"); hdr_font = Font(bold=True, color="FFFFFF")
def style_header(ws, ncol):
    for j in range(1, ncol+1):
        c = ws.cell(1, j); c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

ws = wb.active; ws.title = "Classification"
ws.append(COLS)
for r in rows: ws.append([r[c] for c in COLS])
style_header(ws, len(COLS))
widths=[12,52,18,34,34,14,12,52,11,11]
for j,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w

# QA sheet
qa = wb.create_sheet("QA Summary")
def put(label, val=""):
    qa.append([label, val])
qa.append(["Relationship Life Cycle — Statement Classification QA Summary",""])
qa.cell(1,1).font = Font(bold=True, size=13)
put("")
put("1. Total statements processed", len(rows))
put("")
put("2. Count by Statement_Type","")
for t,_ in by_type.most_common(): put(f"   {t}", by_type[t])
put("")
put("3. Count by Primary Experience Cluster","")
for cid in sorted(by_prim): put(f"   {cid} — {cname.get(cid,'?')}", by_prim[cid])
put("")
put("4. Count marked Behaviorally Actionable", f"{n_action}  ({100*n_action/len(rows):.1f}%)")
put("")
put("5. Count requiring human review (Review_Flag=YES)", f"{n_flag}  ({100*n_flag/len(rows):.1f}%)")
put("   Flag_Category breakdown (multi-category allowed)","")
for cat,_ in flag_cat_counts.most_common(): put(f"   • {cat}", flag_cat_counts[cat])
put("")
put("6. Clusters with unusually few / many statements","")
put("   Unusually FEW (<=12)", ", ".join(f"{c}({by_prim[c]})" for c in sorted(few)) or "none")
put("   Unusually MANY (>=60)", ", ".join(f"{c}({by_prim[c]})" for c in sorted(many)) or "none")
put("   (source-defined bank sizes; not a mapping defect)","")
put("")
put("7. Conceptual overlap (Secondary_EC frequency — may indicate boundary tension)","")
for sec,_ in by_sec.most_common(12): put(f"   → Secondary {sec} — {cname.get(sec,'?')}", by_sec[sec])
qa.column_dimensions["A"].width=58; qa.column_dimensions["B"].width=40
qa.cell(1,1)

# Human Review sheet — validate flags cover exactly the flagged rows
flagged_sids = [r["Statement_ID"] for r in rows if r["Review_Flag"]=="YES"]
mf=[s for s in flagged_sids if s not in flags]; xf=[s for s in flags if s not in set(flagged_sids)]
if mf or xf:
    print("FLAG MISMATCH — missing:", mf, "extra:", xf); sys.exit(1)
VOCAB={"statement-type-ambiguous","self-behavior-boundary","actionability-unclear","relational-condition-vs-experience","self-vs-other-behavior","wording-insufficient","source-data-inconsistency","other"}
for s in flags:
    for c in flags[s]["flag_category"]:
        if c not in VOCAB: print("INVALID CATEGORY:", s, c); sys.exit(1)

hr = wb.create_sheet("Human Review")
HRCOLS = ["Statement_ID","Original_Statement","Primary_EC","RLC_Phase","Statement_Type",
          "Behaviorally_Actionable","Flag_Category","Flag_Reason","Confidence","Secondary_EC","Mapping_Rationale"]
hr.append(HRCOLS)
for r in rows:
    if r["Review_Flag"]!="YES": continue
    f=flags[r["Statement_ID"]]
    hr.append([r["Statement_ID"], r["Original_Statement"], r["Primary_Experience_Cluster"], r["RLC_Phase"],
               r["Statement_Type"], r["Behaviorally_Actionable"], "; ".join(f["flag_category"]),
               f["flag_reason"], r["Confidence"], r["Secondary_Experience_Cluster"], r["Mapping_Rationale"]])
style_header(hr, len(HRCOLS))
hrw=[12,50,30,13,18,13,40,54,11,30,50]
for j,w in enumerate(hrw,1): hr.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w
hr.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HRCOLS))}{hr.max_row}"
ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLS))}{ws.max_row}"
flag_cat_counts=collections.Counter(c for s in flags for c in flags[s]['flag_category'])

out_path = os.path.join(OUT, "RLC_Statement_Classification.xlsx")
wb.save(out_path)
print(f"\nBUILT: {out_path}")
print(f"rows={len(rows)}  actionable={n_action}  flagged={n_flag}")
print("by_type:", dict(by_type))
print("few:", sorted(few), "many:", sorted(many))
print("top secondary:", by_sec.most_common(6))
