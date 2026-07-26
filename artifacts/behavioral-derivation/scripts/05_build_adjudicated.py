#!/usr/bin/env python3
"""Step 5 — validate + build the adjudicated Play-architecture workbook (9 sheets) from ../generated/adjudication/*.json."""
"""Verify + build the adjudicated Play-architecture workbook (6 sheets)."""
import json, glob, os, sys, collections, re
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEN = os.path.join(ROOT,"generated")
OUT = os.path.join(ROOT,"outputs"); os.makedirs(OUT,exist_ok=True)
SP = GEN
ADJ = os.path.join(GEN,"adjudication")
def S(v): return "; ".join(map(str, v)) if isinstance(v, (list, tuple)) else ("" if v is None else str(v))
def IDS(v): return ", ".join(map(str, v)) if isinstance(v, (list, tuple)) else ("" if v is None else str(v))
def L(v): return v if isinstance(v, list) else ([v] if v else [])
bank = {b["sid"]: b for b in json.load(open(f"{SP}/bank.json"))}
beh = {b["sid"]: b for b in json.load(open(f"{SP}/behaviors.json"))}
meta = json.load(open(f"{SP}/cluster_meta.json")); cname = {int(k): v["name"] for k, v in meta.items()}
der = {b["sid"]: b for b in json.load(open(f"{ADJ}/adjudicated_behaviors.json"))}
core = json.load(open(f"{ADJ}/core_plays.json"))
apps = json.load(open(f"{ADJ}/applications.json"))
log = json.load(open(f"{ADJ}/decision_log.json"))
gap = json.load(open(f"{ADJ}/evidence_gap.json")) if os.path.exists(f"{ADJ}/evidence_gap.json") else []
cand = json.load(open(f"{ADJ}/cross_cluster_candidates.json")) if os.path.exists(f"{ADJ}/cross_cluster_candidates.json") else []
deep = json.load(open(f"{ADJ}/cluster1_deepdive.json")) if os.path.exists(f"{ADJ}/cluster1_deepdive.json") else {}
fwc = json.load(open(f"{SP}/fw_competencies.json"))
comp_names = {c["name"].strip() for c in fwc}
GAP = {"Task-supported; no competency cleanly maps", "None available (source gap)", "Task-Supported / Competency Gap", ""}
PF = {"Interrupt","Replace","Increase","Preserve","Observe","Clarify","Decide"}
ES = {"Strongly Supported","Supported with Context Conditions","Task-Supported / Competency Gap","Requires Human Adjudication","Reject"}

# ---- VALIDATE ----
errs = []
if len(der) != 81: errs.append(f"behaviors={len(der)} != 81")
core_ids = {c["core_play_id"] for c in core}
insuff = {s for s in der if der[s]["maintaining_role"] == "Insufficient Evidence"}
for a in apps:
    if a["core_play_id"] not in core_ids: errs.append(f"{a['application_id']}: bad core_play_id")
    if a.get("evidence_status") not in ES: errs.append(f"{a['application_id']}: bad evidence_status {a.get('evidence_status')}")
    for f in L(a.get("play_function")):
        if f not in PF: errs.append(f"{a['application_id']}: bad play_function {f}")
    for sid in L(a.get("behavioral_source_ids")):
        if sid in insuff: errs.append(f"{a['application_id']}: Insufficient sid {sid} as behavioral evidence")
    comp = a.get("supporting_competency", "")
    if comp not in GAP:
        if not any(n.lower() in str(comp).lower() for n in comp_names if n):
            errs.append(f"{a['application_id']}: unknown competency '{comp}'")
app_clusters = {a["experience_cluster"] for a in apps}
for z in (13, 25):
    if z in app_clusters: errs.append(f"cluster {z} should have NO applications")
print("VALIDATION:", "OK ✓" if not errs else f"{len(errs)} ISSUES")
for e in errs[:30]: print("  -", e)
if errs: sys.exit(1)

# ---- BUILD ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
wb = openpyxl.Workbook()
hf = PatternFill("solid", fgColor="1C3557"); hfont = Font(bold=True, color="FFFFFF")
def hdr(ws, n):
    for j in range(1, n + 1):
        c = ws.cell(1, j); c.fill = hf; c.font = hfont; c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(n)}{ws.max_row}"
def widths(ws, ws_widths):
    for j, w in enumerate(ws_widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

# 1. Behavior Derivation — Adjudicated
DCOLS = ["Statement_ID","Experience_Cluster","RLC_Phase","Original_Behavior","Trigger_or_Context","Maintaining_Role",
         "Short_Term_Function","Developmental_Cost","Relevant_Developmental_Task","Relevant_RLC_Competency",
         "Adaptive_Alternative_1","Adaptive_Alternative_2","What_New_Response_Supports","Expected_Discomfort_or_Cost",
         "Context_Conditions","Confidence","Review_Flag"]
ws = wb.active; ws.title = "Behavior Derivation Adjudicated"; ws.append(DCOLS)
for s in sorted(der, key=lambda x: (beh[x]["primary_ec"], x)):
    d = der[s]; b = beh[s]
    ws.append([s, f'{b["primary_ec"]} — {cname.get(b["primary_ec"],"?")}', b["phase"], bank[s]["statement"],
               S(d.get("trigger_or_context")), S(d.get("maintaining_role")), S(d.get("short_term_function")), S(d.get("developmental_cost")),
               S(d.get("relevant_developmental_task")), S(d.get("relevant_rlc_competency")), S(d.get("adaptive_alternative_1")),
               S(d.get("adaptive_alternative_2")), S(d.get("what_new_response_supports")), S(d.get("expected_discomfort_or_cost")),
               S(d.get("context_conditions")), S(d.get("confidence")), "YES" if d.get("review_flag") else "NO"])
hdr(ws, len(DCOLS)); widths(ws, [12,26,13,34,34,22,40,36,20,30,44,40,36,32,34,10,9])

# 2. Core Play Library
napp = collections.Counter(a["core_play_id"] for a in apps)
CCOLS = ["Core_Play_ID","Core_Play_Name_Internal","Behavioral_Target","Play_Function","Canonical_Protocol","Developmental_Tasks","Supporting_Competency","#Applications","Notes"]
cs = wb.create_sheet("Core Play Library"); cs.append(CCOLS)
for c in sorted(core, key=lambda x: x["core_play_id"]):
    cs.append([c["core_play_id"], S(c.get("core_play_name_internal")), S(c.get("behavioral_target")), IDS(c.get("play_function")).replace(","," /"),
               S(c.get("canonical_protocol")), IDS(c.get("developmental_tasks")), S(c.get("supporting_competency")), napp.get(c["core_play_id"],0), S(c.get("notes"))])
hdr(cs, len(CCOLS)); widths(cs, [11,34,40,22,48,26,30,12,40])

# 3. Cluster Play Applications
ACOLS = ["Application_ID","Core_Play_ID","Core_Play_Name_Internal","Experience_Cluster","RLC_Phase","Cluster_Framing","Play_Function",
         "Behavioral_Source_IDs","Contextual_Source_IDs","Supporting_Competency","Evidence_Status","Context_Limitations","Confidence","Review_Flag","Review_Reason"]
cn_by_id = {c["core_play_id"]: c.get("core_play_name_internal","") for c in core}
aS = wb.create_sheet("Cluster Play Applications"); aS.append(ACOLS)
for a in sorted(apps, key=lambda x: (x["core_play_id"], x["experience_cluster"])):
    cid = a["experience_cluster"]
    aS.append([a["application_id"], a["core_play_id"], S(cn_by_id.get(a["core_play_id"])), f'{cid} — {cname.get(cid,"?")}', S(a.get("rlc_phase")),
               S(a.get("cluster_framing")), IDS(a.get("play_function")).replace(","," /"), IDS(a.get("behavioral_source_ids")), IDS(a.get("contextual_source_ids")),
               S(a.get("supporting_competency")), S(a.get("evidence_status")), S(a.get("context_limitations")), S(a.get("confidence")),
               "YES" if a.get("review_flag") else "NO", S(a.get("review_reason"))])
hdr(aS, len(ACOLS)); widths(aS, [15,11,32,24,15,46,20,20,18,30,26,34,10,9,40])

# 4. Human Review
HCOLS = ["Item_Type","ID","Experience_Cluster","Detail","Reason","Status_or_Confidence"]
hrs = wb.create_sheet("Human Review"); hrs.append(HCOLS)
for s in sorted([x for x in der if der[x]["review_flag"]], key=lambda x:(beh[x]["primary_ec"],x)):
    d=der[s]; b=beh[s]
    hrs.append(["Behavior", s, f'{b["primary_ec"]} — {cname.get(b["primary_ec"],"?")}', bank[s]["statement"], S(d.get("review_reason")), S(d.get("confidence"))])
for a in sorted(apps, key=lambda x:(x["core_play_id"],x["experience_cluster"])):
    if a.get("review_flag") or a.get("evidence_status") in ("Requires Human Adjudication","Reject"):
        hrs.append(["Application", a["application_id"], f'{a["experience_cluster"]} — {cname.get(a["experience_cluster"],"?")}',
                    S(cn_by_id.get(a["core_play_id"])), S(a.get("review_reason")) or S(a.get("evidence_status")), S(a.get("evidence_status"))])
hdr(hrs, len(HCOLS)); widths(hrs, [12,15,26,40,50,26])

# 5. Global QA — Resolved
qa = wb.create_sheet("Global QA — Resolved"); qa.cell(1,1,"Script 2 — Play-Layer Adjudication: Global QA (Resolved)").font=Font(bold=True,size=13)
r=[3]
def line(a,b=""):
    qa.cell(r[0],1,a); qa.cell(r[0],2,b); r[0]+=1
es_dist = collections.Counter(a["evidence_status"] for a in apps)
line("Core Plays", len(core)); line("Cluster Applications", len(apps)); line("Decision-log entries", len(log))
line("Zero-Play clusters preserved", "13, 25")
line("")
line("Evidence_Status distribution",""); [line(f"   {k}", es_dist[k]) for k in ES if es_dist[k]]
line("")
line("RESOLUTION OF ORIGINAL QA FLAGS","")
line("  same_behavior_diff_function", "RESOLVED — STM-0055/1085 competency normalized to Emotional Intimacy → Gradual Self-Disclosure (CP-06); STM-0446 c14 vs c7 governed by CP-13 (competency where phase supports, gap Cross-Phase).")
line("  inconsistent_alternatives", "RESOLVED — STM-0449 adopts the canonical CP-06 'safe person + watch' protocol (was generic).")
line("  generic_advice", "RESOLVED — STM-0449 + c14 keep-the-peace framing rewritten to conditional CP-06 protocol with observation; reciprocity-pacing kept (competency-anchored, conditional).")
line("  unsupported_mechanisms", "RESOLVED — STM-0052/0782 'name the fear' → behavioral; STM-0357 function → behavioral; CP-20/0655/0661 'out of fear' → 'even though part of you knows it isn't good'.")
line("  over_consolidated_plays", "RESOLVED via SPLIT — c5 Play → CP-01 (pace pursuit) + CP-02 (don't dismiss available); c11 Play → CP-04 (see-by-date) + CP-05 (name what happened); c14 mega-Play → CP-06/CP-13/CP-14/CP-15.")
line("  duplicative_plays_across_clusters", "RESOLVED via NORMALIZATION into Core Plays — testing (CP-08 c3/c6), guard/disclosure (CP-06 c3/c27), imbalance (CP-12 c7/c10), reach-first (CP-01 c2/c15), see-by-date (CP-04 c5/c11/c14/c24).")
line("")
line("MECHANICAL RE-CHECKS (post-adjudication)","")
bad_ev=[(a['application_id'],sid) for a in apps for sid in L(a.get('behavioral_source_ids')) if sid in insuff]
line("  Insufficient-Evidence used as behavioral evidence", str(bad_ev) if bad_ev else "none ✓")
unknown=[a['application_id'] for a in apps if a.get('supporting_competency') not in GAP and not any(n.lower() in str(a.get('supporting_competency')).lower() for n in comp_names if n)]
line("  Invented/unknown competency labels", ", ".join(unknown) or "none ✓")
line("  Applications requiring human adjudication", es_dist.get("Requires Human Adjudication",0))
line("  Task-Supported / Competency Gap (source-gap phases)", es_dist.get("Task-Supported / Competency Gap",0))
qa.column_dimensions["A"].width=52; qa.column_dimensions["B"].width=110

# 6. Decision Log
LCOLS = ["#","Change_Type","Target","Description"]
ls = wb.create_sheet("Decision Log"); ls.append(LCOLS)
for i,e in enumerate(log,1): ls.append([i, S(e.get("change_type")), S(e.get("target")), S(e.get("description"))])
hdr(ls, len(LCOLS)); widths(ls, [5,22,30,100])

# 7. Evidence Gap Analysis
if gap:
    GCOLS=["Cluster","Name","Total_Statements","Self_Behavior_Stmts","Behaviorally_Actionable","Approved_Applications","Unresolved_Applications","Classification","Rationale"]
    gs=wb.create_sheet("Evidence Gap Analysis"); gs.append(GCOLS)
    for g in gap:
        gs.append([g["cluster"], g["name"], g["total_statements"], g["self_behavior"], g["actionable"], g["approved_apps"], g["unresolved_apps"], g["classification"], g["rationale"]])
    hdr(gs, len(GCOLS)); widths(gs, [8,44,14,14,14,14,14,28,70])

# 8. Cross-Cluster Candidates
if cand:
    XCOLS=["Target_Experience_Cluster","Candidate_Statement_ID","Candidate_Statement","Canonical_Source_Cluster","RLC_Phase","Why_It_May_Be_Relevant","Relevant_Developmental_Task","Relevant_RLC_Competency_if_any","Evidence_Link_Type","Confidence","Human_Review_Required"]
    xs=wb.create_sheet("Cross-Cluster Candidates"); xs.append(XCOLS)
    for c in cand:
        xs.append([c["target_cluster"], c["candidate_statement_id"], c["candidate_statement"], c["canonical_source_cluster"], c["rlc_phase"],
                   c["why_it_may_be_relevant"], c["relevant_developmental_task"], c["relevant_rlc_competency_if_any"], c["evidence_link_type"], c["confidence"], c["human_review_required"]])
    hdr(xs, len(XCOLS)); widths(xs, [30,20,40,30,14,66,22,30,26,11,16])

# 9. Cluster 1 Deep Dive
if deep:
    ds=wb.create_sheet("Cluster 1 Deep Dive"); ds.append(["Section","Content"])
    LABELS={"A_phenomenology":"A. What C1 statements establish (phenomenology)","B_own_behaviors":"B. Behaviors directly established by C1's own Self-Behaviors","C_task":"C. What the Exploration task (Discernment) requires","D_competencies":"D. Relevant approved Exploration competencies","E_candidates_elsewhere":"E. Behavioral statements elsewhere potentially relevant","F_can_they_connect":"F. Can they connect without an unsupported mechanism?"}
    for k in ("A_phenomenology","B_own_behaviors","C_task","D_competencies","E_candidates_elsewhere","F_can_they_connect"):
        ds.append([LABELS[k], deep.get(k,"")])
    hdr(ds, 2); widths(ds, [46,120]);
    for row in ds.iter_rows(min_row=2): row[1].alignment=Alignment(wrap_text=True, vertical="top")

out=os.path.join(OUT,"RLC_Play_Architecture_Adjudicated.xlsx"); wb.save(out)
print("BUILT:", out)
print("core plays:", len(core), "| applications:", len(apps), "| decisions:", len(log))
print("evidence status:", dict(es_dist))
print("flagged behaviors:", sum(1 for s in der if der[s]['review_flag']), "| flagged/adjudicate apps:", sum(1 for a in apps if a.get('review_flag') or a.get('evidence_status') in ('Requires Human Adjudication','Reject')))
